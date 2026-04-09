import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import os
import platform
import tempfile
import io
import math
import unicodedata
from datetime import datetime, date
import fitz  # PyMuPDF
from PIL import Image, ImageTk, ImageDraw, ImageFont
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill

# ─── Unicode Font Detection ───────────────────────────────────────────────────
def find_unicode_font():
    system = platform.system()
    if system == "Windows":
        paths = [
            "C:/Windows/Fonts/seguiemj.ttf", "C:/Windows/Fonts/segoeuisym.ttf",
            "C:/Windows/Fonts/seguisym.ttf",  "C:/Windows/Fonts/segoeui.ttf",
            "C:/Windows/Fonts/msyh.ttc",      "C:/Windows/Fonts/msyhbd.ttc",
            "C:/Windows/Fonts/simhei.ttf",    "C:/Windows/Fonts/arialuni.ttf",
            "C:/Windows/Fonts/cambria.ttc",   "C:/Windows/Fonts/arial.ttf",
            "C:/Windows/Fonts/calibri.ttf",
        ]
    elif system == "Darwin":
        paths = [
            "/System/Library/Fonts/Apple Color Emoji.ttc",
            "/System/Library/Fonts/Apple Symbols.ttf",
            "/Library/Fonts/Arial Unicode MS.ttf",
            "/System/Library/Fonts/Helvetica.ttc",
        ]
    else:
        paths = [
            "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
            "/usr/share/fonts/truetype/noto/NotoSansSymbols2-Regular.ttf",
            "/usr/share/fonts/truetype/noto/NotoColorEmoji.ttf",
            "/usr/share/fonts/truetype/noto/NotoSansMath-Regular.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
            "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
            "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
            "/usr/share/fonts/truetype/stix/STIXGeneral.ttf",
        ]
    for p in paths:
        if os.path.exists(p):
            return p
    return None

UNICODE_FONT_PATH = find_unicode_font()


class PDFMarkerApp:
    # ═══════════════════════════════════════════════════════════════════════
    # INIT
    # ═══════════════════════════════════════════════════════════════════════
    def __init__(self, root):
        self.root = root
        self.root.title("PDF 标记工具")
        self.root.geometry("1500x920")
        self.root.configure(bg="#ececec")
        style = ttk.Style()
        style.configure("TNotebook.Tab", padding=[12, 4])
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        self.mark_frame   = ttk.Frame(self.notebook)
        self.apply_frame  = ttk.Frame(self.notebook)
        self.verify_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.mark_frame,   text=" 📌 标记 PDF ")
        self.notebook.add(self.apply_frame,  text=" 📄 应用标记 ")
        self.notebook.add(self.verify_frame, text=" ✅ 核对 Excel ")
        self._setup_mark_tab()
        self._setup_apply_tab()
        self._setup_verify_tab()

    # ═══════════════════════════════════════════════════════════════════════
    # STATIC HELPERS — 3-Point Affine Transform
    # ═══════════════════════════════════════════════════════════════════════
    @staticmethod
    def _transform_point(lx, ly, orig_refs, new_refs):
        """
        Transform local coords (relative to orig p1) → new PDF absolute coords.
        Degrades gracefully:
          0 refs   → identity (return original absolute)
          1 ref    → pure translation
          2 refs   → similarity (translation + rotation + uniform scale)
          3+ refs  → full affine (translation + rotation + non-uniform scale + shear)
        orig_refs / new_refs : list of (x,y) tuples, or None
        lx, ly : marker offset relative to orig p1 = m["x"], m["y"]
        """
        if not new_refs:
            if orig_refs:
                return orig_refs[0][0] + lx, orig_refs[0][1] + ly
            return lx, ly

        nx1, ny1 = new_refs[0]
        n = min(len(orig_refs) if orig_refs else 0, len(new_refs))

        if n < 2:
            return nx1 + lx, ny1 + ly

        ox1, oy1 = orig_refs[0]
        ox2, oy2 = orig_refs[1]
        nx2, ny2 = new_refs[1]

        if n == 2:
            # ── 2-point: similarity transform ──────────────────────────
            dox = ox2-ox1;  doy = oy2-oy1
            dnx = nx2-nx1;  dny = ny2-ny1
            len_o = math.sqrt(dox*dox + doy*doy)
            len_n = math.sqrt(dnx*dnx + dny*dny)
            if len_o < 1e-6:
                return nx1+lx, ny1+ly
            s = len_n / len_o
            a = math.atan2(dny, dnx) - math.atan2(doy, dox)
            ca, sa = math.cos(a), math.sin(a)
            return (lx*ca - ly*sa)*s + nx1, (lx*sa + ly*ca)*s + ny1

        # ── 3-point: full affine transform ─────────────────────────────
        ox3, oy3 = orig_refs[2]
        nx3, ny3 = new_refs[2]
        # basis vectors in orig space (relative to p1)
        u1x, u1y = ox2-ox1, oy2-oy1
        u2x, u2y = ox3-ox1, oy3-oy1
        # basis vectors in new space (relative to q1)
        v1x, v1y = nx2-nx1, ny2-ny1
        v2x, v2y = nx3-nx1, ny3-ny1
        # solve A * [u1|u2] = [v1|v2]   →   A = [v1|v2] * [u1|u2]^-1
        det_U = u1x*u2y - u2x*u1y
        if abs(det_U) < 1e-6:
            # degenerate (collinear refs) → fall back to 2-point similarity
            dox = ox2-ox1;  doy = oy2-oy1
            dnx = nx2-nx1;  dny = ny2-ny1
            len_o = math.sqrt(dox*dox+doy*doy)
            len_n = math.sqrt(dnx*dnx+dny*dny)
            if len_o < 1e-6:
                return nx1+lx, ny1+ly
            s = len_n/len_o
            a = math.atan2(dny,dnx) - math.atan2(doy,dox)
            ca,sa = math.cos(a),math.sin(a)
            return (lx*ca-ly*sa)*s+nx1, (lx*sa+ly*ca)*s+ny1
        inv = 1.0 / det_U
        a00 = ( v1x*u2y - v2x*u1y) * inv
        a01 = (-v1x*u2x + v2x*u1x) * inv
        a10 = ( v1y*u2y - v2y*u1y) * inv
        a11 = (-v1y*u2x + v2y*u1x) * inv
        return a00*lx + a01*ly + nx1, a10*lx + a11*ly + ny1

    @staticmethod
    def _transform_rect(lx, ly, lw, lh, orig_refs, new_refs):
        """
        Transform an axis-aligned rectangle through the affine map.
        Returns (new_x, new_y, new_w, new_h) as axis-aligned bounding box of
        the four transformed corners.
        """
        tp = PDFMarkerApp._transform_point
        corners = [
            tp(lx,    ly,    orig_refs, new_refs),
            tp(lx+lw, ly,    orig_refs, new_refs),
            tp(lx,    ly+lh, orig_refs, new_refs),
            tp(lx+lw, ly+lh, orig_refs, new_refs),
        ]
        xs = [p[0] for p in corners]; ys = [p[1] for p in corners]
        nx = min(xs); ny = min(ys)
        return nx, ny, max(xs)-nx, max(ys)-ny


    # ═══════════════════════════════════════════════════════════════════════
    # TAB 1 — MARK
    # ═══════════════════════════════════════════════════════════════════════
    def _setup_mark_tab(self):
        self.mark_pdf_path   = None
        self.mark_doc        = None
        self.mark_page_index = 0
        self.mark_zoom       = 1.5
        self.mark_photo      = None
        self.mark_sx         = 1.0
        self.mark_sy         = 1.0
        # per-page: up to 3 ref points [(x,y), ...]
        self.mark_ref_points = {}
        self.markers         = []
        self.mark_mode       = "idle"   # "set_ref1/2/3" | "marking"
        self.drag_start      = None
        self.rubber_band     = None

        r1 = ttk.Frame(self.mark_frame)
        r1.pack(fill=tk.X, padx=10, pady=(8,2))
        ttk.Button(r1, text="📂 选择 PDF", command=self._mark_open_pdf).pack(side=tk.LEFT, padx=3)
        self.lbl_mark_pdf = ttk.Label(r1, text="未选择文件", foreground="gray")
        self.lbl_mark_pdf.pack(side=tk.LEFT, padx=6)
        ttk.Separator(r1, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)
        self.btn_m_ref1   = ttk.Button(r1, text="⊕ 参考点1", command=self._mark_enter_ref1_mode, state=tk.DISABLED)
        self.btn_m_ref1.pack(side=tk.LEFT, padx=3)
        self.btn_m_ref2   = ttk.Button(r1, text="⊕ 参考点2", command=self._mark_enter_ref2_mode, state=tk.DISABLED)
        self.btn_m_ref2.pack(side=tk.LEFT, padx=3)
        self.btn_m_ref3   = ttk.Button(r1, text="⊕ 参考点3", command=self._mark_enter_ref3_mode, state=tk.DISABLED)
        self.btn_m_ref3.pack(side=tk.LEFT, padx=3)
        self.btn_m_start  = ttk.Button(r1, text="▶ 开始标记",  command=self._mark_start,  state=tk.DISABLED)
        self.btn_m_start.pack(side=tk.LEFT, padx=3)
        self.btn_m_finish = ttk.Button(r1, text="✔ 完成标记",  command=self._mark_finish, state=tk.DISABLED)
        self.btn_m_finish.pack(side=tk.LEFT, padx=3)
        self.btn_m_undo   = ttk.Button(r1, text="↩ 撤销",      command=self._mark_undo,   state=tk.DISABLED)
        self.btn_m_undo.pack(side=tk.LEFT, padx=3)
        self.btn_m_clear  = ttk.Button(r1, text="🗑 清空",      command=self._mark_clear,  state=tk.DISABLED)
        self.btn_m_clear.pack(side=tk.LEFT, padx=3)

        r2 = ttk.Frame(self.mark_frame)
        r2.pack(fill=tk.X, padx=10, pady=2)
        ttk.Button(r2, text="◀", width=3, command=self._mark_prev_page).pack(side=tk.LEFT)
        self.lbl_mark_page = ttk.Label(r2, text=" — / — ")
        self.lbl_mark_page.pack(side=tk.LEFT)
        ttk.Button(r2, text="▶", width=3, command=self._mark_next_page).pack(side=tk.LEFT)
        self.lbl_mark_status = ttk.Label(r2, text="请先选择 PDF 文件", foreground="#2471a3")
        self.lbl_mark_status.pack(side=tk.LEFT, padx=20)

        cf = ttk.Frame(self.mark_frame)
        cf.pack(fill=tk.BOTH, expand=True, padx=10, pady=6)
        self.mark_canvas = tk.Canvas(cf, bg="#404040", cursor="crosshair", highlightthickness=0)
        self.mark_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb = ttk.Scrollbar(cf, orient=tk.VERTICAL, command=self.mark_canvas.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb = ttk.Scrollbar(self.mark_frame, orient=tk.HORIZONTAL, command=self.mark_canvas.xview)
        hsb.pack(fill=tk.X, padx=10, pady=(0,4))
        self.mark_canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.mark_canvas.bind("<ButtonPress-1>",  self._mc_press)
        self.mark_canvas.bind("<B1-Motion>",       self._mc_motion)
        self.mark_canvas.bind("<ButtonRelease-1>", self._mc_release)

    # ── Mark: file ─────────────────────────────────────────────────────
    def _mark_open_pdf(self):
        path = filedialog.askopenfilename(filetypes=[("PDF 文件","*.pdf")])
        if not path: return
        self.mark_pdf_path = path
        self.mark_doc      = fitz.open(path)
        self.mark_page_index = 0
        self.markers.clear()
        self.mark_ref_points.clear()
        self.lbl_mark_pdf.config(text=os.path.basename(path), foreground="black")
        self.mark_mode = "idle"
        self._update_mark_controls()
        self._mark_render()
        self._mark_status("PDF 已加载 → 请为每页依次设置三个参考点，再开始标记")

    def _mark_prev_page(self):
        if self.mark_doc and self.mark_page_index > 0:
            self.mark_page_index -= 1; self.mark_mode = "idle"; self._mark_render()

    def _mark_next_page(self):
        if self.mark_doc and self.mark_page_index < len(self.mark_doc)-1:
            self.mark_page_index += 1; self.mark_mode = "idle"; self._mark_render()

    # ── Mark: ref-point modes ──────────────────────────────────────────
    def _mark_enter_ref1_mode(self):
        if not self.mark_doc: return
        self.mark_mode = "set_ref1"
        self.mark_canvas.config(cursor="tcross")
        self._mark_status(f"🎯 参考点1：点击第 {self.mark_page_index+1} 页固定特征位置（如已知角点）")

    def _mark_enter_ref2_mode(self):
        if not self.mark_doc: return
        if not self.mark_ref_points.get(self.mark_page_index):
            messagebox.showwarning("提示","请先设置参考点1！"); return
        self.mark_mode = "set_ref2"
        self.mark_canvas.config(cursor="tcross")
        self._mark_status(f"🎯 参考点2：点击第 {self.mark_page_index+1} 页第二个固定特征位置")

    def _mark_enter_ref3_mode(self):
        if not self.mark_doc: return
        pts = self.mark_ref_points.get(self.mark_page_index, [])
        if len(pts) < 2:
            messagebox.showwarning("提示","请先设置参考点1和参考点2！"); return
        self.mark_mode = "set_ref3"
        self.mark_canvas.config(cursor="tcross")
        self._mark_status(f"🎯 参考点3：点击第 {self.mark_page_index+1} 页第三个固定特征位置（不与前两点共线）")

    def _mark_start(self):
        if not self.mark_doc: return
        pts = self.mark_ref_points.get(self.mark_page_index, [])
        if len(pts) < 3:
            messagebox.showwarning("提示", f"请先为第 {self.mark_page_index+1} 页设置三个参考点（当前已设 {len(pts)} 个）！")
            return
        self.mark_mode = "marking"
        self.mark_canvas.config(cursor="crosshair")
        self._update_mark_controls()
        self._mark_status(f"✏ 标记模式：第 {self.mark_page_index+1} 页 — 拖拽画出矩形区域")

    def _mark_undo(self):
        if self.markers:
            self.markers.pop(); self._mark_render()
            self._mark_status(f"已撤销，当前共 {len(self.markers)} 个标记")

    def _mark_clear(self):
        self.markers.clear(); self._mark_render(); self._mark_status("已清空所有标记")

    # ── Mark: canvas events ────────────────────────────────────────────
    def _mc_press(self, event):
        if not self.mark_doc: return
        cx = self.mark_canvas.canvasx(event.x)
        cy = self.mark_canvas.canvasy(event.y)
        px, py = cx/self.mark_sx, cy/self.mark_sy

        if self.mark_mode in ("set_ref1","set_ref2","set_ref3"):
            idx = int(self.mark_mode[-1]) - 1   # 0/1/2
            pts = list(self.mark_ref_points.get(self.mark_page_index, []))
            # minimum distance check
            for i, ep in enumerate(pts):
                if i != idx:
                    if math.sqrt((px-ep[0])**2+(py-ep[1])**2) < 20:
                        self._mark_status("⚠ 新参考点与已有参考点距离太近（< 20pt），请重新选择"); return
            while len(pts) < idx: pts.append((0,0))
            if len(pts) <= idx: pts.append((px,py))
            else: pts[idx] = (px,py)
            self.mark_ref_points[self.mark_page_index] = pts
            self.mark_mode = "idle"
            self.mark_canvas.config(cursor="crosshair")
            self._mark_render()
            self._update_mark_controls()
            names = ["参考点1","参考点2","参考点3"]
            remaining = 3 - len(pts)
            hint = f"→ 还需设置 {remaining} 个参考点" if remaining > 0 else "→ 三点参考系就绪，可以开始标记"
            self._mark_status(f"✅ {names[idx]} ({px:.1f},{py:.1f}) 已设置 {hint}")

        elif self.mark_mode == "marking":
            self.drag_start  = (cx, cy)
            self.rubber_band = self.mark_canvas.create_rectangle(
                cx, cy, cx, cy, outline="#e74c3c", width=2, dash=(5,3), tags="rubber")

    def _mc_motion(self, event):
        if self.mark_mode != "marking" or not self.drag_start: return
        cx = self.mark_canvas.canvasx(event.x); cy = self.mark_canvas.canvasy(event.y)
        self.mark_canvas.coords(self.rubber_band, *self.drag_start, cx, cy)

    def _mc_release(self, event):
        if self.mark_mode != "marking" or not self.drag_start: return
        cx = self.mark_canvas.canvasx(event.x); cy = self.mark_canvas.canvasy(event.y)
        x0,y0 = self.drag_start; x1,y1 = cx,cy
        self.mark_canvas.delete("rubber"); self.rubber_band = None; self.drag_start = None
        if abs(x1-x0)<6 or abs(y1-y0)<6: return
        if x0>x1: x0,x1=x1,x0
        if y0>y1: y0,y1=y1,y0
        pts = self.mark_ref_points.get(self.mark_page_index,[])
        if len(pts)<3:
            messagebox.showwarning("提示",f"第 {self.mark_page_index+1} 页参考点不足（需要3个）"); return
        bx,by = pts[0]
        self.markers.append({"id":len(self.markers)+1, "page":self.mark_page_index,
                              "x":round(x0/self.mark_sx-bx,3), "y":round(y0/self.mark_sy-by,3),
                              "width":round((x1-x0)/self.mark_sx,3), "height":round((y1-y0)/self.mark_sy,3)})
        self._mark_render()
        self._mark_status(f"已添加标记 {len(self.markers)}")

    # ── Mark: render ───────────────────────────────────────────────────
    REF_COLS  = ["#f0b429","#27ae60","#e74c3c"]   # yellow, green, red
    REF_NAMES = ["R1","R2","R3"]

    def _mark_render(self):
        if not self.mark_doc:
            self.mark_canvas.delete("all"); self.lbl_mark_page.config(text=" — / — "); return
        page = self.mark_doc[self.mark_page_index]
        pix  = page.get_pixmap(matrix=fitz.Matrix(self.mark_zoom,self.mark_zoom), alpha=False)
        img  = Image.frombytes("RGB",[pix.width,pix.height],pix.samples)
        self.mark_photo = ImageTk.PhotoImage(img)
        pr = page.rect
        self.mark_sx = pix.width/pr.width; self.mark_sy = pix.height/pr.height
        self.mark_canvas.config(scrollregion=(0,0,pix.width,pix.height))
        self.mark_canvas.delete("all")
        self.mark_canvas.create_image(0,0,anchor=tk.NW,image=self.mark_photo)
        self.lbl_mark_page.config(text=f" {self.mark_page_index+1} / {len(self.mark_doc)} ")
        # Draw ref triangle
        pts = self.mark_ref_points.get(self.mark_page_index,[])
        screen_pts = [(p[0]*self.mark_sx, p[1]*self.mark_sy) for p in pts]
        if len(screen_pts)==3:
            self.mark_canvas.create_polygon(
                *[c for sp in screen_pts for c in sp],
                outline="#95a5a6", fill="", width=1, dash=(4,4), tags="refpts")
        elif len(screen_pts)==2:
            self.mark_canvas.create_line(*screen_pts[0],*screen_pts[1],
                                         fill="#95a5a6",width=1,dash=(4,4),tags="refpts")
        for i,(bx,by) in enumerate(screen_pts):
            col=self.REF_COLS[i]; r=9
            self.mark_canvas.create_oval(bx-r,by-r,bx+r,by+r,outline=col,width=2,tags="refpts")
            self.mark_canvas.create_line(bx-r-5,by,bx+r+5,by,fill=col,width=2,tags="refpts")
            self.mark_canvas.create_line(bx,by-r-5,bx,by+r+5,fill=col,width=2,tags="refpts")
            self.mark_canvas.create_text(bx+r+8,by,text=self.REF_NAMES[i],
                                         fill=col,font=("Arial",9,"bold"),tags="refpts")
        # Draw markers
        ref_p1 = pts[0] if pts else (0,0)
        for m in self.markers:
            if m["page"]!=self.mark_page_index: continue
            sx=(ref_p1[0]+m["x"])*self.mark_sx; sy=(ref_p1[1]+m["y"])*self.mark_sy
            sw=m["width"]*self.mark_sx;          sh=m["height"]*self.mark_sy
            self._draw_marker_on_canvas(self.mark_canvas,sx,sy,sw,sh,m["id"],"#e74c3c")
        self._update_mark_controls()

    def _update_mark_controls(self):
        if not self.mark_doc:
            for b in (self.btn_m_ref1,self.btn_m_ref2,self.btn_m_ref3,
                      self.btn_m_start,self.btn_m_finish,self.btn_m_undo,self.btn_m_clear):
                b.config(state=tk.DISABLED); return
        pts = self.mark_ref_points.get(self.mark_page_index,[])
        n = len(pts)
        self.btn_m_ref1.config(state=tk.NORMAL)
        self.btn_m_ref2.config(state=tk.NORMAL if n>=1 else tk.DISABLED)
        self.btn_m_ref3.config(state=tk.NORMAL if n>=2 else tk.DISABLED)
        if self.mark_mode=="marking":
            self.btn_m_start.config(state=tk.DISABLED)
            self.btn_m_finish.config(state=tk.NORMAL)
        else:
            self.btn_m_start.config(state=tk.NORMAL if n>=3 else tk.DISABLED)
            self.btn_m_finish.config(state=tk.DISABLED)
        has_m = len(self.markers)>0
        self.btn_m_undo.config(state=tk.NORMAL if has_m else tk.DISABLED)
        self.btn_m_clear.config(state=tk.NORMAL if has_m else tk.DISABLED)

    # ── Mark: save ─────────────────────────────────────────────────────
    def _mark_finish(self):
        if not self.markers:
            messagebox.showwarning("提示","还没有任何标记！"); return
        save_dir = filedialog.askdirectory(title="选择保存目录")
        if not save_dir: return
        base = os.path.splitext(os.path.basename(self.mark_pdf_path))[0]
        rp_data = {}
        for pidx,pts in self.mark_ref_points.items():
            entry = {}
            for i,pt in enumerate(pts[:3]):
                entry[f"p{i+1}"] = {"x":round(pt[0],3),"y":round(pt[1],3)}
            rp_data[str(pidx)] = entry
        data = {"ref_points":rp_data, "markers":self.markers}
        json_path = os.path.join(save_dir,f"{base}_markers.json")
        with open(json_path,"w",encoding="utf-8") as f:
            json.dump(data,f,ensure_ascii=False,indent=2)
        xlsx_path = os.path.join(save_dir,f"{base}_markers.xlsx")
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="标记信息"
        ws.append(["序号","坐标信息（相对参考点1）","内容（第3列）","截图（第4列）"])
        for cell in ws[1]:
            cell.font=Font(bold=True,name="Arial")
            cell.fill=PatternFill("solid",fgColor="D6EAF8")
            cell.alignment=Alignment(horizontal="center")
        for m in self.markers:
            coord=f"x={m['x']}, y={m['y']}, w={m['width']}, h={m['height']}, page={m['page']+1}"
            ws.append([m["id"],coord,"",""])
        ws.column_dimensions["A"].width=8; ws.column_dimensions["B"].width=52
        ws.column_dimensions["C"].width=24; ws.column_dimensions["D"].width=28
        wb.save(xlsx_path)
        self.mark_mode="idle"; self._update_mark_controls()
        messagebox.showinfo("保存成功",f"✔ {len(self.markers)} 个标记已保存：\nJSON: {json_path}\nExcel: {xlsx_path}")
        self._mark_status(f"✔ 已保存 {len(self.markers)} 个标记")

    def _mark_status(self, text): self.lbl_mark_status.config(text=text)


    # ═══════════════════════════════════════════════════════════════════════
    # TAB 2 — APPLY
    # ═══════════════════════════════════════════════════════════════════════
    def _setup_apply_tab(self):
        self.apply_pdf_path   = None
        self.apply_json_path  = None
        self.apply_xlsx_path  = None
        self.apply_doc        = None
        self.apply_markers    = []
        self.apply_orig_refs  = {}   # from JSON:  page → [(ox1,oy1),(ox2,oy2),(ox3,oy3)]
        self.apply_new_refs   = {}   # user-set:   page → growing list of up to 3 pts
        self.apply_excel_data = []
        self.apply_page_index = 0
        self.apply_zoom       = 1.5
        self.apply_photo      = None
        self.apply_sx         = 1.0; self.apply_sy = 1.0
        self.apply_mode       = "idle"
        self._tmp_files       = []
        self.apply_selected_id= None
        self.apply_font_sizes = {}

        r1 = ttk.Frame(self.apply_frame)
        r1.pack(fill=tk.X, padx=10, pady=(8,2))
        ttk.Button(r1,text="📂 选择 PDF",      command=self._apply_open_pdf).pack(side=tk.LEFT,padx=3)
        self.lbl_apply_pdf  = ttk.Label(r1,text="未选择",foreground="gray"); self.lbl_apply_pdf.pack(side=tk.LEFT,padx=4)
        ttk.Separator(r1,orient=tk.VERTICAL).pack(side=tk.LEFT,fill=tk.Y,padx=8)
        ttk.Button(r1,text="📋 选择标记 JSON", command=self._apply_open_json).pack(side=tk.LEFT,padx=3)
        self.lbl_apply_json = ttk.Label(r1,text="未选择",foreground="gray"); self.lbl_apply_json.pack(side=tk.LEFT,padx=4)
        ttk.Separator(r1,orient=tk.VERTICAL).pack(side=tk.LEFT,fill=tk.Y,padx=8)
        ttk.Button(r1,text="📊 选择 Excel（可选）",command=self._apply_open_xlsx).pack(side=tk.LEFT,padx=3)
        self.lbl_apply_xlsx = ttk.Label(r1,text="未选择",foreground="gray"); self.lbl_apply_xlsx.pack(side=tk.LEFT,padx=4)

        r2 = ttk.Frame(self.apply_frame)
        r2.pack(fill=tk.X, padx=10, pady=2)
        self.btn_a_ref1 = ttk.Button(r2,text="⊕ 参考点1",command=self._apply_enter_ref1_mode,state=tk.DISABLED)
        self.btn_a_ref1.pack(side=tk.LEFT,padx=3)
        self.btn_a_ref2 = ttk.Button(r2,text="⊕ 参考点2",command=self._apply_enter_ref2_mode,state=tk.DISABLED)
        self.btn_a_ref2.pack(side=tk.LEFT,padx=3)
        self.btn_a_ref3 = ttk.Button(r2,text="⊕ 参考点3",command=self._apply_enter_ref3_mode,state=tk.DISABLED)
        self.btn_a_ref3.pack(side=tk.LEFT,padx=3)
        self.lbl_a_refstatus = ttk.Label(r2,text="未设置参考点",foreground="gray")
        self.lbl_a_refstatus.pack(side=tk.LEFT,padx=6)
        ttk.Separator(r2,orient=tk.VERTICAL).pack(side=tk.LEFT,fill=tk.Y,padx=8)
        ttk.Button(r2,text="◀",width=3,command=self._apply_prev_page).pack(side=tk.LEFT)
        self.lbl_apply_page = ttk.Label(r2,text=" — / — "); self.lbl_apply_page.pack(side=tk.LEFT)
        ttk.Button(r2,text="▶",width=3,command=self._apply_next_page).pack(side=tk.LEFT)
        ttk.Separator(r2,orient=tk.VERTICAL).pack(side=tk.LEFT,fill=tk.Y,padx=8)
        ttk.Button(r2,text="💾 导出带标记 PDF",     command=self._apply_export_pdf).pack(side=tk.LEFT,padx=3)
        ttk.Button(r2,text="📸 截图存入 Excel 第4列",command=self._apply_screenshot_to_excel).pack(side=tk.LEFT,padx=3)
        self.lbl_apply_status = ttk.Label(r2,text="请选择 PDF 和标记 JSON",foreground="#2471a3")
        self.lbl_apply_status.pack(side=tk.LEFT,padx=14)

        main_pane = ttk.Panedwindow(self.apply_frame,orient=tk.HORIZONTAL)
        main_pane.pack(fill=tk.BOTH,expand=True,padx=10,pady=6)
        left  = ttk.Frame(main_pane)
        right = ttk.Frame(main_pane,width=420)
        main_pane.add(left,weight=4); main_pane.add(right,weight=1)

        cf = ttk.Frame(left); cf.pack(fill=tk.BOTH,expand=True)
        self.apply_canvas = tk.Canvas(cf,bg="#404040",highlightthickness=0)
        self.apply_canvas.pack(side=tk.LEFT,fill=tk.BOTH,expand=True)
        vsb=ttk.Scrollbar(cf,orient=tk.VERTICAL,command=self.apply_canvas.yview); vsb.pack(side=tk.RIGHT,fill=tk.Y)
        hsb=ttk.Scrollbar(left,orient=tk.HORIZONTAL,command=self.apply_canvas.xview); hsb.pack(fill=tk.X,pady=(0,4))
        self.apply_canvas.configure(yscrollcommand=vsb.set,xscrollcommand=hsb.set)
        self.apply_canvas.bind("<ButtonPress-1>",self._ac_press)

        ttk.Label(right,text="第3列手动修正",font=("Arial",11,"bold")).pack(anchor="w",pady=(0,4))
        top_tools=ttk.Frame(right); top_tools.pack(fill=tk.X,pady=(0,6))
        for txt,cmd in [("刷新列表",self._refresh_apply_tree),("清空当前",self._apply_clear_selected),
                        ("删除当前",self._apply_delete_selected),("应用修改",self._apply_save_selected)]:
            ttk.Button(top_tools,text=txt,command=cmd).pack(side=tk.LEFT,padx=2)

        tree_frame=ttk.Frame(right); tree_frame.pack(fill=tk.BOTH,expand=False,pady=(0,6))
        cols=("id","page","preview")
        self.apply_tree=ttk.Treeview(tree_frame,columns=cols,show="headings",height=12,selectmode="browse")
        self.apply_tree.heading("id",text="序号"); self.apply_tree.heading("page",text="页")
        self.apply_tree.heading("preview",text="第3列内容预览")
        self.apply_tree.column("id",width=50,anchor="center"); self.apply_tree.column("page",width=45,anchor="center")
        self.apply_tree.column("preview",width=250,anchor="w")
        ysb2=ttk.Scrollbar(tree_frame,orient=tk.VERTICAL,command=self.apply_tree.yview)
        self.apply_tree.configure(yscrollcommand=ysb2.set)
        self.apply_tree.pack(side=tk.LEFT,fill=tk.BOTH,expand=True); ysb2.pack(side=tk.RIGHT,fill=tk.Y)
        self.apply_tree.bind("<<TreeviewSelect>>",self._on_apply_tree_select)
        self.apply_tree.bind("<Button-1>",self._on_tree_click)

        edit_box=ttk.LabelFrame(right,text="内容编辑"); edit_box.pack(fill=tk.BOTH,expand=True)
        self.apply_text=tk.Text(edit_box,height=7,wrap="word"); self.apply_text.pack(fill=tk.BOTH,expand=True,padx=4,pady=4)
        sb1=ttk.Frame(edit_box); sb1.pack(fill=tk.X,padx=4,pady=(0,4))
        sb2=ttk.Frame(edit_box); sb2.pack(fill=tk.X,padx=4,pady=(0,4))
        for i,(name,sym) in enumerate([("空心圆","○"),("圈杠","⊘"),("空框","☐"),("勾框","☑"),
                                        ("叉框","☒"),("打勾","✓"),("粗勾","✔"),("实心圆","●")]):
            ttk.Button(sb1 if i<4 else sb2,text=name,width=8,
                       command=lambda s=sym: self._insert_symbol(s)).pack(side=tk.LEFT,padx=2,pady=1)
        ff=ttk.Frame(edit_box); ff.pack(fill=tk.X,padx=4,pady=(0,4))
        ttk.Label(ff,text="字号:").pack(side=tk.LEFT)
        self.apply_font_spin=ttk.Spinbox(ff,from_=4,to=72,width=6,command=self._apply_font_size_changed)
        self.apply_font_spin.pack(side=tk.LEFT,padx=2)
        ttk.Label(ff,text="pt").pack(side=tk.LEFT)
        bt=ttk.Frame(edit_box); bt.pack(fill=tk.X,pady=(4,0))
        for txt,cmd in [("← 载入到编辑框",self._load_selected_to_editor),("保存到当前项",self._apply_save_selected),
                        ("上一项",self._select_prev_tree_item),("下一项",self._select_next_tree_item)]:
            ttk.Button(bt,text=txt,command=cmd).pack(side=tk.LEFT,padx=2)

    # ── Apply: file openers ─────────────────────────────────────────────
    def _apply_open_pdf(self):
        path=filedialog.askopenfilename(filetypes=[("PDF 文件","*.pdf")])
        if not path: return
        self.apply_pdf_path=path; self.apply_doc=fitz.open(path)
        self.apply_page_index=0; self.apply_new_refs.clear()
        self.lbl_apply_pdf.config(text=os.path.basename(path),foreground="black")
        self._apply_refresh_ref_btn(); self._apply_render(); self._apply_status("PDF 已加载")

    def _apply_open_json(self):
        path=filedialog.askopenfilename(filetypes=[("JSON 文件","*.json")])
        if not path: return
        with open(path,"r",encoding="utf-8") as f: data=json.load(f)
        self.apply_markers=data.get("markers",[])
        self.apply_orig_refs={}
        if "ref_points" in data:
            for k,v in data["ref_points"].items():
                try:
                    pts=[]
                    for pkey in ("p1","p2","p3"):
                        if pkey in v: pts.append((float(v[pkey]["x"]),float(v[pkey]["y"])))
                    if pts: self.apply_orig_refs[int(k)]=pts
                except: pass
        elif "baselines" in data and isinstance(data["baselines"],dict):
            for k,v in data["baselines"].items():
                try: self.apply_orig_refs[int(k)]=[(float(v["x"]),float(v["y"]))]
                except: pass
        elif "baseline" in data:
            try: self.apply_orig_refs[0]=[(float(data["baseline"]["x"]),float(data["baseline"]["y"]))]
            except: pass
        self.apply_font_sizes={m["id"]:12 for m in self.apply_markers if "id" in m}
        self.apply_json_path=path
        self.lbl_apply_json.config(text=os.path.basename(path),foreground="black")
        self._apply_refresh_ref_btn(); self._refresh_apply_tree(); self._apply_render()
        self._apply_status(f"JSON 已加载，{len(self.apply_markers)} 个标记")

    def _apply_open_xlsx(self):
        path=filedialog.askopenfilename(filetypes=[("Excel 文件","*.xlsx *.xls")])
        if not path: return
        wb=openpyxl.load_workbook(path,data_only=False); ws=wb.active
        self.apply_excel_data=[]
        for row in ws.iter_rows(min_row=2):
            cell=row[2] if len(row)>=3 else None
            self.apply_excel_data.append(self._cell_display_text(cell) if cell else "")
        self.apply_xlsx_path=path
        self.lbl_apply_xlsx.config(text=os.path.basename(path),foreground="black")
        self._refresh_apply_tree(); self._apply_render()
        self._apply_status(f"Excel 已加载，读取第3列 {len(self.apply_excel_data)} 行")

    def _apply_refresh_ref_btn(self):
        ok=bool(self.apply_doc and self.apply_markers)
        self.btn_a_ref1.config(state=tk.NORMAL if ok else tk.DISABLED)
        self.btn_a_ref2.config(state=tk.DISABLED)
        self.btn_a_ref3.config(state=tk.DISABLED)

    # ── Apply: ref-point modes ──────────────────────────────────────────
    def _apply_enter_ref1_mode(self):
        if not self.apply_doc: return
        self.apply_mode="set_ref1"; self.apply_canvas.config(cursor="tcross")
        self._apply_status(f"🎯 参考点1：点击第 {self.apply_page_index+1} 页对应的参考点1位置")

    def _apply_enter_ref2_mode(self):
        if not self.apply_doc: return
        if not self.apply_new_refs.get(self.apply_page_index):
            messagebox.showwarning("提示","请先设置参考点1！"); return
        self.apply_mode="set_ref2"; self.apply_canvas.config(cursor="tcross")
        self._apply_status(f"🎯 参考点2：点击第 {self.apply_page_index+1} 页对应的参考点2位置")

    def _apply_enter_ref3_mode(self):
        if not self.apply_doc: return
        pts=self.apply_new_refs.get(self.apply_page_index,[])
        if len(pts)<2:
            messagebox.showwarning("提示","请先设置参考点1和参考点2！"); return
        self.apply_mode="set_ref3"; self.apply_canvas.config(cursor="tcross")
        self._apply_status(f"🎯 参考点3：点击第 {self.apply_page_index+1} 页对应的参考点3位置（不与前两点共线）")

    def _ac_press(self, event):
        if self.apply_mode not in ("set_ref1","set_ref2","set_ref3") or not self.apply_doc: return
        cx=self.apply_canvas.canvasx(event.x); cy=self.apply_canvas.canvasy(event.y)
        px,py=cx/self.apply_sx, cy/self.apply_sy
        idx=int(self.apply_mode[-1])-1   # 0/1/2
        pts=list(self.apply_new_refs.get(self.apply_page_index,[]))
        for i,ep in enumerate(pts):
            if i!=idx and math.sqrt((px-ep[0])**2+(py-ep[1])**2)<10:
                self._apply_status("⚠ 参考点太近，请选择更远位置"); return
        while len(pts)<idx: pts.append((0,0))
        if len(pts)<=idx: pts.append((px,py))
        else: pts[idx]=(px,py)
        self.apply_new_refs[self.apply_page_index]=pts
        self.apply_mode="idle"; self.apply_canvas.config(cursor="arrow")
        # unlock next button
        if idx==0: self.btn_a_ref2.config(state=tk.NORMAL)
        elif idx==1: self.btn_a_ref3.config(state=tk.NORMAL)
        self._apply_render()
        names=["参考点1","参考点2","参考点3"]
        mode_hints=["请继续设置参考点2","请继续设置参考点3","三点仿射变换已启用"]
        self._apply_status(f"✅ {names[idx]} ({px:.1f},{py:.1f}) → {mode_hints[idx]}")

    # ── Apply: marker positioning ──────────────────────────────────────
    def _marker_abs_pdf(self, m):
        pidx=m["page"]
        return self._transform_point(m["x"],m["y"],
                                      self.apply_orig_refs.get(pidx),
                                      self.apply_new_refs.get(pidx))

    # ── Apply: compose image ────────────────────────────────────────────
    def _compose_page_image(self, page_index, zoom=None):
        if not self.apply_doc: return None,1.0,1.0
        z=self.apply_zoom if zoom is None else zoom
        page=self.apply_doc[page_index]
        pix=page.get_pixmap(matrix=fitz.Matrix(z,z),alpha=False)
        base=Image.frombytes("RGB",[pix.width,pix.height],pix.samples).convert("RGBA")
        pr=page.rect; sx=pix.width/pr.width; sy=pix.height/pr.height
        draw=ImageDraw.Draw(base,"RGBA")
        # Draw ref points + triangle
        orig_refs=self.apply_orig_refs.get(page_index)
        new_refs =self.apply_new_refs.get(page_index)
        active   =new_refs if new_refs else orig_refs
        REF_COLS_PIL=[(240,180,41,255),(39,174,96,255),(231,76,60,255)]
        if active:
            sp=[(p[0]*sx,p[1]*sy) for p in active]
            if len(sp)==3:
                draw.polygon([sp[0],sp[1],sp[2]],outline=(149,165,166,140),fill=None)
                # fix for Pillow polygon outline
                for i in range(3):
                    draw.line([sp[i],sp[(i+1)%3]],fill=(149,165,166,140),width=1)
            elif len(sp)==2:
                draw.line([sp[0],sp[1]],fill=(149,165,166,140),width=1)
            for i,(bx,by) in enumerate(sp):
                col=REF_COLS_PIL[i]; r=9
                draw.ellipse((bx-r,by-r,bx+r,by+r),outline=col,width=2)
                draw.line((bx-r-5,by,bx+r+5,by),fill=col,width=2)
                draw.line((bx,by-r-5,bx,by+r+5),fill=col,width=2)
        badge_font=self._load_font(9)
        for m in self.apply_markers:
            if m.get("page")!=page_index: continue
            orig_r=self.apply_orig_refs.get(page_index)
            new_r =self.apply_new_refs.get(page_index)
            nx,ny,nw,nh=self._transform_rect(m["x"],m["y"],m["width"],m["height"],orig_r,new_r)
            x=int(round(nx*sx)); y=int(round(ny*sy))
            w=max(1,int(round(nw*sx))); h=max(1,int(round(nh*sy)))
            mid=m["id"]
            draw.rectangle((x,y,x+w,y+h),outline=(41,128,185,255),width=2)
            badge_w=max(22,len(str(mid))*7+6)
            draw.rectangle((x,y-18,x+badge_w,y),fill=(41,128,185,255),outline=(41,128,185,255))
            if badge_font:
                tb=draw.textbbox((0,0),str(mid),font=badge_font)
                tw2,th2=tb[2]-tb[0],tb[3]-tb[1]
                draw.text((x+(badge_w-tw2)/2,y-18+(18-th2)/2-tb[1]),str(mid),font=badge_font,fill=(255,255,255,255))
            else:
                draw.text((x+4,y-16),str(mid),fill=(255,255,255,255))
            idx2=mid-1
            if self.apply_excel_data and 0<=idx2<len(self.apply_excel_data):
                txt=self.apply_excel_data[idx2]
                if txt:
                    fsize=self.apply_font_sizes.get(mid,12)
                    txt_img=self._make_text_image(txt,max(12,w-2),max(12,h-2),fixed_font_size=fsize)
                    base.alpha_composite(txt_img,(x+1,y+1))
        return base.convert("RGB"),sx,sy

    def _apply_render(self):
        if not self.apply_doc:
            self.apply_canvas.delete("all"); self.lbl_apply_page.config(text=" — / — ")
            self.lbl_a_refstatus.config(text="未设置参考点",foreground="gray"); return
        composed,sx,sy=self._compose_page_image(self.apply_page_index,zoom=self.apply_zoom)
        if composed is None: return
        self.apply_sx=sx; self.apply_sy=sy
        self.apply_photo=ImageTk.PhotoImage(composed)
        w,h=composed.size
        self.apply_canvas.config(scrollregion=(0,0,w,h))
        self.apply_canvas.delete("all")
        self.apply_canvas.create_image(0,0,anchor=tk.NW,image=self.apply_photo)
        self.lbl_apply_page.config(text=f" {self.apply_page_index+1} / {len(self.apply_doc)} ")
        new_refs=self.apply_new_refs.get(self.apply_page_index,[])
        n=len(new_refs)
        if n>=3:   self.lbl_a_refstatus.config(text="✅ 三点仿射变换已启用",foreground="#1a5276")
        elif n==2: self.lbl_a_refstatus.config(text="⚠ 两点相似变换（建议再设第3点）",foreground="#8b6914")
        elif n==1: self.lbl_a_refstatus.config(text="⚠ 仅参考点1（纯平移）",foreground="#8b6914")
        else:
            orig=self.apply_orig_refs.get(self.apply_page_index)
            self.lbl_a_refstatus.config(
                text=f"ℹ 使用JSON参考点（{len(orig)}个）" if orig else "未设置参考点",
                foreground="gray")
        self._refresh_apply_tree()

    def _apply_prev_page(self):
        if self.apply_doc and self.apply_page_index>0:
            self.apply_page_index-=1; self.apply_mode="idle"; self.apply_canvas.config(cursor="arrow"); self._apply_render()

    def _apply_next_page(self):
        if self.apply_doc and self.apply_page_index<len(self.apply_doc)-1:
            self.apply_page_index+=1; self.apply_mode="idle"; self.apply_canvas.config(cursor="arrow"); self._apply_render()

    def _apply_status(self,text): self.lbl_apply_status.config(text=text)


    # ── Apply: tree / editor ─────────────────────────────────────────────
    def _refresh_apply_tree(self):
        if not hasattr(self,"apply_tree"): return
        for item in self.apply_tree.get_children(): self.apply_tree.delete(item)
        for m in sorted(self.apply_markers,key=lambda x:x.get("id",0)):
            mid=m.get("id",0); idx=mid-1
            text=self.apply_excel_data[idx] if 0<=idx<len(self.apply_excel_data) else ""
            prev=text.replace("\n"," ⏎ ").strip()
            if len(prev)>40: prev=prev[:40]+"…"
            self.apply_tree.insert("","end",iid=str(mid),values=(mid,m.get("page",0)+1,prev))
        if self.apply_selected_id and self.apply_tree.exists(str(self.apply_selected_id)):
            self.apply_tree.selection_set(str(self.apply_selected_id))
            self.apply_tree.see(str(self.apply_selected_id))

    def _on_apply_tree_select(self,event=None):
        sel=self.apply_tree.selection()
        if not sel: return
        try: self.apply_selected_id=int(sel[0])
        except: return
        self._load_selected_to_editor()

    def _load_selected_to_editor(self):
        if self.apply_selected_id is None: return
        idx=self.apply_selected_id-1
        text=self.apply_excel_data[idx] if 0<=idx<len(self.apply_excel_data) else ""
        self.apply_text.delete("1.0",tk.END); self.apply_text.insert("1.0",text)
        self.apply_font_spin.set(self.apply_font_sizes.get(self.apply_selected_id,12))

    def _insert_symbol(self,sym):
        try: self.apply_text.insert(tk.INSERT,sym)
        except: pass

    def _selected_tree_id(self):
        sel=self.apply_tree.selection()
        if sel:
            try: return int(sel[0])
            except: pass
        return self.apply_selected_id

    def _apply_save_selected(self):
        mid=self._selected_tree_id()
        if not mid: messagebox.showwarning("提示","请先在右侧列表中选中一个标记"); return
        content=self.apply_text.get("1.0",tk.END).rstrip("\n")
        idx=mid-1
        while len(self.apply_excel_data)<=idx: self.apply_excel_data.append("")
        self.apply_excel_data[idx]=content
        try: self.apply_font_sizes[mid]=int(self.apply_font_spin.get())
        except: pass
        self.apply_selected_id=mid
        self._refresh_apply_tree(); self._apply_render(); self._apply_status(f"已更新第 {mid} 项内容")

    def _apply_clear_selected(self): self.apply_text.delete("1.0",tk.END)

    def _apply_delete_selected(self):
        mid=self._selected_tree_id()
        if not mid: return
        if messagebox.askyesno("确认",f"确定清空第 {mid} 项内容吗？"):
            idx=mid-1
            while len(self.apply_excel_data)<=idx: self.apply_excel_data.append("")
            self.apply_excel_data[idx]=""
            self.apply_text.delete("1.0",tk.END)
            self._refresh_apply_tree(); self._apply_render(); self._apply_status(f"已清空第 {mid} 项内容")

    def _select_prev_tree_item(self):
        items=self.apply_tree.get_children()
        if not items: return
        ids=[int(i) for i in items]; cur=self.apply_selected_id
        iid=str(ids[max(0,ids.index(cur)-1)]) if cur in ids else str(ids[0])
        self.apply_tree.selection_set(iid); self.apply_tree.see(iid); self._on_apply_tree_select()

    def _select_next_tree_item(self):
        items=self.apply_tree.get_children()
        if not items: return
        ids=[int(i) for i in items]; cur=self.apply_selected_id
        iid=str(ids[min(len(ids)-1,ids.index(cur)+1)]) if cur in ids else str(ids[0])
        self.apply_tree.selection_set(iid); self.apply_tree.see(iid); self._on_apply_tree_select()

    def _apply_font_size_changed(self):
        if self.apply_selected_id is None: return
        try:
            self.apply_font_sizes[self.apply_selected_id]=int(self.apply_font_spin.get())
            self._apply_render()
        except ValueError: pass

    def _on_tree_click(self,event):
        if self.apply_tree.identify_column(event.x)=="#1":
            item=self.apply_tree.identify_row(event.y)
            if item:
                try:
                    mid=int(item)
                    for m in self.apply_markers:
                        if m.get("id")==mid:
                            page=m.get("page",0)
                            if self.apply_page_index!=page:
                                self.apply_page_index=page; self.apply_mode="idle"
                                self.apply_canvas.config(cursor="arrow"); self._apply_render()
                            self.apply_tree.selection_set(item); self.apply_tree.see(item)
                            self._on_apply_tree_select(); break
                except: pass

    # ── Apply: export ───────────────────────────────────────────────────
    def _apply_export_pdf(self):
        if not self.apply_doc or not self.apply_markers:
            messagebox.showwarning("提示","请先加载 PDF 和 JSON"); return
        save_path=filedialog.asksaveasfilename(defaultextension=".pdf",filetypes=[("PDF 文件","*.pdf")])
        if not save_path: return
        out=fitz.open()
        for pidx in range(len(self.apply_doc)):
            composed,_,_=self._compose_page_image(pidx,zoom=self.apply_zoom)
            if composed is None: continue
            buf=io.BytesIO(); composed.save(buf,format="PNG")
            pg=out.new_page(width=composed.width,height=composed.height)
            pg.insert_image(pg.rect,stream=buf.getvalue(),keep_proportion=False,overlay=True)
        out.save(save_path); out.close()
        messagebox.showinfo("导出成功",f"带标记 PDF 已保存：\n{save_path}")

    def _apply_screenshot_to_excel(self):
        if not self.apply_doc or not self.apply_markers:
            messagebox.showwarning("提示","请先加载 PDF 和 JSON"); return
        xlsx_src=filedialog.askopenfilename(title="选择要更新的 Excel 文件",filetypes=[("Excel 文件","*.xlsx")])
        if not xlsx_src: return
        wb=openpyxl.load_workbook(xlsx_src); ws=wb.active
        ZOOM=3.0; page_imgs={}
        for m in self.apply_markers:
            pidx=m["page"]
            if pidx not in page_imgs:
                composed,sx,sy=self._compose_page_image(pidx,zoom=ZOOM)
                page_imgs[pidx]=(composed,sx,sy)
        for m in self.apply_markers:
            pidx=m["page"]; img,sx,sy=page_imgs[pidx]
            orig_r=self.apply_orig_refs.get(pidx); new_r=self.apply_new_refs.get(pidx)
            nx,ny,nw,nh=self._transform_rect(m["x"],m["y"],m["width"],m["height"],orig_r,new_r)
            x0=max(0,int(nx*sx)); y0=max(0,int(ny*sy))
            x1=min(img.width,int((nx+nw)*sx)); y1=min(img.height,int((ny+nh)*sy))
            if x1<=x0 or y1<=y0: continue
            cropped=img.crop((x0,y0,x1,y1))
            ImageDraw.Draw(cropped).rectangle([0,0,cropped.width-1,cropped.height-1],outline=(41,128,185),width=3)
            tmp=tempfile.NamedTemporaryFile(suffix=".png",delete=False); tmp.close()
            cropped.save(tmp.name,"PNG"); self._tmp_files.append(tmp.name)
            row_num=m["id"]+1
            while ws.max_row<row_num: ws.append(["","","",""])
            xl_img=XLImage(tmp.name); ow,oh=cropped.size; sc=min(80/oh,240/ow)
            xl_img.width=int(ow*sc); xl_img.height=int(oh*sc); xl_img.anchor=f"D{row_num}"
            ws.add_image(xl_img); ws.row_dimensions[row_num].height=xl_img.height*0.75+4
        ws.column_dimensions["D"].width=34
        save_path=filedialog.asksaveasfilename(title="保存更新后的 Excel",defaultextension=".xlsx",
                   filetypes=[("Excel 文件","*.xlsx")],initialfile=os.path.basename(xlsx_src))
        if save_path:
            wb.save(save_path); messagebox.showinfo("完成",f"截图已写入 Excel 第4列：\n{save_path}")
        for p in self._tmp_files:
            try: os.unlink(p)
            except: pass
        self._tmp_files.clear()


    # ═══════════════════════════════════════════════════════════════════════
    # TAB 3 — VERIFY EXCEL
    # ═══════════════════════════════════════════════════════════════════════
    def _setup_verify_tab(self):
        self.verify_pdf_path   = None
        self.verify_json_path  = None
        self.verify_xlsx_path  = None
        self.verify_doc        = None
        self.verify_markers    = []
        self.verify_orig_refs  = {}   # page → [(x1,y1),(x2,y2),(x3,y3)]
        self.verify_excel_data = []
        self.verify_statuses   = {}   # {mid: "correct"/"incorrect"/"unknown"}
        self.verify_notes      = {}   # {mid: str}
        self.verify_page_index = 0
        self.verify_zoom       = 1.5
        self.verify_photo      = None
        self.verify_sx         = 1.0; self.verify_sy = 1.0
        self.verify_selected_id= None
        self.verify_status_var = tk.StringVar(value="unknown")

        r1=ttk.Frame(self.verify_frame); r1.pack(fill=tk.X,padx=10,pady=(8,2))
        ttk.Button(r1,text="📂 选择 PDF",       command=self._verify_open_pdf).pack(side=tk.LEFT,padx=3)
        self.lbl_verify_pdf=ttk.Label(r1,text="未选择",foreground="gray"); self.lbl_verify_pdf.pack(side=tk.LEFT,padx=4)
        ttk.Separator(r1,orient=tk.VERTICAL).pack(side=tk.LEFT,fill=tk.Y,padx=8)
        ttk.Button(r1,text="📋 选择标记 JSON",  command=self._verify_open_json).pack(side=tk.LEFT,padx=3)
        self.lbl_verify_json=ttk.Label(r1,text="未选择",foreground="gray"); self.lbl_verify_json.pack(side=tk.LEFT,padx=4)
        ttk.Separator(r1,orient=tk.VERTICAL).pack(side=tk.LEFT,fill=tk.Y,padx=8)
        ttk.Button(r1,text="📊 导入 Excel 核对数据",command=self._verify_open_xlsx).pack(side=tk.LEFT,padx=3)
        self.lbl_verify_xlsx=ttk.Label(r1,text="未选择",foreground="gray"); self.lbl_verify_xlsx.pack(side=tk.LEFT,padx=4)
        ttk.Separator(r1,orient=tk.VERTICAL).pack(side=tk.LEFT,fill=tk.Y,padx=8)
        ttk.Button(r1,text="💾 导出核对结果",   command=self._verify_export_results).pack(side=tk.LEFT,padx=3)

        r2=ttk.Frame(self.verify_frame); r2.pack(fill=tk.X,padx=10,pady=2)
        ttk.Button(r2,text="◀",width=3,command=self._verify_prev_page).pack(side=tk.LEFT)
        self.lbl_verify_page=ttk.Label(r2,text=" — / — "); self.lbl_verify_page.pack(side=tk.LEFT)
        ttk.Button(r2,text="▶",width=3,command=self._verify_next_page).pack(side=tk.LEFT)
        ttk.Separator(r2,orient=tk.VERTICAL).pack(side=tk.LEFT,fill=tk.Y,padx=8)
        self.lbl_verify_summary=ttk.Label(r2,text="总计: 0 | ✓ 正确: 0 | ✗ 错误: 0 | ? 未核对: 0",foreground="#2471a3")
        self.lbl_verify_summary.pack(side=tk.LEFT,padx=6)

        main_pane=ttk.Panedwindow(self.verify_frame,orient=tk.HORIZONTAL)
        main_pane.pack(fill=tk.BOTH,expand=True,padx=10,pady=6)
        left=ttk.Frame(main_pane); right=ttk.Frame(main_pane,width=460)
        main_pane.add(left,weight=3); main_pane.add(right,weight=2)

        cf=ttk.Frame(left); cf.pack(fill=tk.BOTH,expand=True)
        self.verify_canvas=tk.Canvas(cf,bg="#404040",highlightthickness=0)
        self.verify_canvas.pack(side=tk.LEFT,fill=tk.BOTH,expand=True)
        vsb=ttk.Scrollbar(cf,orient=tk.VERTICAL,command=self.verify_canvas.yview); vsb.pack(side=tk.RIGHT,fill=tk.Y)
        hsb=ttk.Scrollbar(left,orient=tk.HORIZONTAL,command=self.verify_canvas.xview); hsb.pack(fill=tk.X,pady=(0,4))
        self.verify_canvas.configure(yscrollcommand=vsb.set,xscrollcommand=hsb.set)
        self.verify_canvas.bind("<ButtonPress-1>",self._vc_press)

        ttk.Label(right,text="Excel 数据核对",font=("Arial",11,"bold")).pack(anchor="w",pady=(0,4))
        tree_frame=ttk.Frame(right); tree_frame.pack(fill=tk.BOTH,expand=True,pady=(0,4))
        cols=("id","content","status")
        self.verify_tree=ttk.Treeview(tree_frame,columns=cols,show="headings",height=14,selectmode="browse")
        self.verify_tree.heading("id",text="#"); self.verify_tree.heading("content",text="Excel 内容预览")
        self.verify_tree.heading("status",text="状态")
        self.verify_tree.column("id",width=40,anchor="center"); self.verify_tree.column("content",width=270,anchor="w")
        self.verify_tree.column("status",width=80,anchor="center")
        self.verify_tree.tag_configure("correct",  background="#d5f5e3",foreground="#1d8348")
        self.verify_tree.tag_configure("incorrect",background="#fadbd8",foreground="#922b21")
        self.verify_tree.tag_configure("unknown",  background="#fef9e7",foreground="#7d6608")
        ysb2=ttk.Scrollbar(tree_frame,orient=tk.VERTICAL,command=self.verify_tree.yview)
        self.verify_tree.configure(yscrollcommand=ysb2.set)
        self.verify_tree.pack(side=tk.LEFT,fill=tk.BOTH,expand=True); ysb2.pack(side=tk.RIGHT,fill=tk.Y)
        self.verify_tree.bind("<<TreeviewSelect>>",self._on_verify_tree_select)

        ctrl=ttk.LabelFrame(right,text="核对操作"); ctrl.pack(fill=tk.X,pady=(4,4))
        self.lbl_verify_detail=ttk.Label(ctrl,text="请在上方列表选择一行",foreground="gray")
        self.lbl_verify_detail.pack(anchor="w",padx=6,pady=(6,2))
        cf2=ttk.LabelFrame(ctrl,text="Excel 内容"); cf2.pack(fill=tk.X,padx=6,pady=(2,4))
        self.verify_content_text=tk.Text(cf2,height=4,wrap="word",state=tk.DISABLED,bg="#f8f9fa")
        self.verify_content_text.pack(fill=tk.X,padx=4,pady=4)
        rr=ttk.Frame(ctrl); rr.pack(fill=tk.X,padx=6,pady=(0,4))
        ttk.Label(rr,text="核对结果：",font=("Arial",10,"bold")).pack(side=tk.LEFT)
        ttk.Radiobutton(rr,text="✓ 正确", variable=self.verify_status_var,value="correct",
                        command=self._verify_save_status).pack(side=tk.LEFT,padx=8)
        ttk.Radiobutton(rr,text="✗ 错误", variable=self.verify_status_var,value="incorrect",
                        command=self._verify_save_status).pack(side=tk.LEFT,padx=8)
        ttk.Radiobutton(rr,text="? 未核对",variable=self.verify_status_var,value="unknown",
                        command=self._verify_save_status).pack(side=tk.LEFT,padx=8)
        nr=ttk.Frame(ctrl); nr.pack(fill=tk.X,padx=6,pady=(0,4))
        ttk.Label(nr,text="备注：").pack(side=tk.LEFT)
        self.verify_notes_var=tk.StringVar()
        self.verify_notes_entry=ttk.Entry(nr,textvariable=self.verify_notes_var,width=32)
        self.verify_notes_entry.pack(side=tk.LEFT,padx=4,fill=tk.X,expand=True)
        ttk.Button(nr,text="保存备注",command=self._verify_save_notes).pack(side=tk.LEFT,padx=4)
        navr=ttk.Frame(ctrl); navr.pack(fill=tk.X,padx=6,pady=(0,6))
        for txt,cmd in [("◀ 上一项",self._verify_prev_item),("下一项 ▶",self._verify_next_item),
                        ("定位到PDF",self._verify_locate_in_pdf)]:
            ttk.Button(navr,text=txt,command=cmd).pack(side=tk.LEFT,padx=4)

    # ── Verify: file openers ─────────────────────────────────────────────
    def _verify_open_pdf(self):
        path=filedialog.askopenfilename(filetypes=[("PDF 文件","*.pdf")])
        if not path: return
        self.verify_pdf_path=path; self.verify_doc=fitz.open(path); self.verify_page_index=0
        self.lbl_verify_pdf.config(text=os.path.basename(path),foreground="black"); self._verify_render()

    def _verify_open_json(self):
        path=filedialog.askopenfilename(filetypes=[("JSON 文件","*.json")])
        if not path: return
        with open(path,"r",encoding="utf-8") as f: data=json.load(f)
        self.verify_markers=data.get("markers",[])
        self.verify_orig_refs={}
        if "ref_points" in data:
            for k,v in data["ref_points"].items():
                try:
                    pts=[]
                    for pkey in ("p1","p2","p3"):
                        if pkey in v: pts.append((float(v[pkey]["x"]),float(v[pkey]["y"])))
                    if pts: self.verify_orig_refs[int(k)]=pts
                except: pass
        elif "baselines" in data and isinstance(data["baselines"],dict):
            for k,v in data["baselines"].items():
                try: self.verify_orig_refs[int(k)]=[(float(v["x"]),float(v["y"]))]
                except: pass
        for m in self.verify_markers:
            mid=m.get("id")
            if mid and mid not in self.verify_statuses: self.verify_statuses[mid]="unknown"
        self.verify_json_path=path
        self.lbl_verify_json.config(text=os.path.basename(path),foreground="black")
        self._refresh_verify_tree(); self._verify_render(); self._update_verify_summary()

    def _verify_open_xlsx(self):
        path=filedialog.askopenfilename(filetypes=[("Excel 文件","*.xlsx *.xls")])
        if not path: return
        wb=openpyxl.load_workbook(path,data_only=False); ws=wb.active
        self.verify_excel_data=[]
        for row in ws.iter_rows(min_row=2):
            row_text=" | ".join(self._cell_display_text(c) for c in row if c and self._cell_display_text(c))
            self.verify_excel_data.append(row_text)
        self.verify_xlsx_path=path
        self.lbl_verify_xlsx.config(text=os.path.basename(path),foreground="black")
        self._refresh_verify_tree(); self._verify_render(); self._update_verify_summary()

    # ── Verify: tree ─────────────────────────────────────────────────────
    def _refresh_verify_tree(self):
        if not hasattr(self,"verify_tree"): return
        for item in self.verify_tree.get_children(): self.verify_tree.delete(item)
        for m in sorted(self.verify_markers,key=lambda x:x.get("id",0)):
            mid=m.get("id",0); idx=mid-1
            text=self.verify_excel_data[idx] if 0<=idx<len(self.verify_excel_data) else "（无数据）"
            prev=text.replace("\n"," ").strip()
            if len(prev)>42: prev=prev[:42]+"…"
            status=self.verify_statuses.get(mid,"unknown")
            icon={"correct":"✓","incorrect":"✗","unknown":"?"}[status]
            self.verify_tree.insert("","end",iid=str(mid),values=(mid,prev,icon),tags=(status,))
        if self.verify_selected_id and self.verify_tree.exists(str(self.verify_selected_id)):
            self.verify_tree.selection_set(str(self.verify_selected_id))
            self.verify_tree.see(str(self.verify_selected_id))

    def _on_verify_tree_select(self,event=None):
        sel=self.verify_tree.selection()
        if not sel: return
        try: mid=int(sel[0])
        except: return
        self.verify_selected_id=mid
        self._load_verify_detail(mid)
        for m in self.verify_markers:
            if m.get("id")==mid:
                page=m.get("page",0)
                if self.verify_page_index!=page:
                    self.verify_page_index=page; self._verify_render()
                else:
                    self._verify_render()
                self._verify_scroll_to_marker(m); break

    def _load_verify_detail(self,mid):
        idx=mid-1
        text=self.verify_excel_data[idx] if 0<=idx<len(self.verify_excel_data) else ""
        self.lbl_verify_detail.config(text=f"标记 #{mid}",foreground="#1a5276")
        self.verify_content_text.config(state=tk.NORMAL)
        self.verify_content_text.delete("1.0",tk.END)
        self.verify_content_text.insert("1.0",text)
        self.verify_content_text.config(state=tk.DISABLED)
        self.verify_status_var.set(self.verify_statuses.get(mid,"unknown"))
        self.verify_notes_var.set(self.verify_notes.get(mid,""))

    def _vc_press(self,event):
        if not self.verify_doc or not self.verify_markers: return
        cx=self.verify_canvas.canvasx(event.x); cy=self.verify_canvas.canvasy(event.y)
        for m in self.verify_markers:
            if m.get("page",0)!=self.verify_page_index: continue
            refs=self.verify_orig_refs.get(m["page"])
            nx,ny,nw,nh=self._transform_rect(m["x"],m["y"],m["width"],m["height"],refs,None)
            x=nx*self.verify_sx; y=ny*self.verify_sy
            w=nw*self.verify_sx; h=nh*self.verify_sy
            if x<=cx<=x+w and y<=cy<=y+h:
                self.verify_selected_id=m["id"]
                if self.verify_tree.exists(str(m["id"])):
                    self.verify_tree.selection_set(str(m["id"])); self.verify_tree.see(str(m["id"]))
                self._load_verify_detail(m["id"]); self._verify_render(); return

    def _verify_save_status(self):
        if self.verify_selected_id is None: return
        self.verify_statuses[self.verify_selected_id]=self.verify_status_var.get()
        self._refresh_verify_tree(); self._verify_render(); self._update_verify_summary()

    def _verify_save_notes(self):
        if self.verify_selected_id is None: return
        self.verify_notes[self.verify_selected_id]=self.verify_notes_var.get()

    def _verify_prev_item(self):
        items=self.verify_tree.get_children()
        if not items: return
        ids=[int(i) for i in items]; cur=self.verify_selected_id
        iid=str(ids[max(0,ids.index(cur)-1)]) if cur in ids else str(ids[0])
        self.verify_tree.selection_set(iid); self.verify_tree.see(iid); self._on_verify_tree_select()

    def _verify_next_item(self):
        items=self.verify_tree.get_children()
        if not items: return
        ids=[int(i) for i in items]; cur=self.verify_selected_id
        iid=str(ids[min(len(ids)-1,ids.index(cur)+1)]) if cur in ids else str(ids[0])
        self.verify_tree.selection_set(iid); self.verify_tree.see(iid); self._on_verify_tree_select()

    def _verify_locate_in_pdf(self):
        if self.verify_selected_id is None: return
        for m in self.verify_markers:
            if m.get("id")==self.verify_selected_id:
                if self.verify_page_index!=m.get("page",0):
                    self.verify_page_index=m.get("page",0); self._verify_render()
                self._verify_scroll_to_marker(m); return

    def _verify_scroll_to_marker(self,m):
        refs=self.verify_orig_refs.get(m["page"])
        nx,ny,nw,nh=self._transform_rect(m["x"],m["y"],m["width"],m["height"],refs,None)
        cx=(nx+nw/2)*self.verify_sx; cy=(ny+nh/2)*self.verify_sy
        sr=self.verify_canvas.cget("scrollregion")
        if not sr: return
        try: parts=str(sr).split(); tw=float(parts[2]); th=float(parts[3])
        except: return
        cw=self.verify_canvas.winfo_width(); ch=self.verify_canvas.winfo_height()
        self.verify_canvas.xview_moveto(max(0.0,min(1.0,(cx-cw/2)/tw)))
        self.verify_canvas.yview_moveto(max(0.0,min(1.0,(cy-ch/2)/th)))

    def _verify_prev_page(self):
        if self.verify_doc and self.verify_page_index>0:
            self.verify_page_index-=1; self._verify_render()

    def _verify_next_page(self):
        if self.verify_doc and self.verify_page_index<len(self.verify_doc)-1:
            self.verify_page_index+=1; self._verify_render()

    # ── Verify: render ───────────────────────────────────────────────────
    def _verify_render(self):
        if not self.verify_doc: self.verify_canvas.delete("all"); self.lbl_verify_page.config(text=" — / — "); return
        pidx=self.verify_page_index; page=self.verify_doc[pidx]
        pix=page.get_pixmap(matrix=fitz.Matrix(self.verify_zoom,self.verify_zoom),alpha=False)
        base=Image.frombytes("RGB",[pix.width,pix.height],pix.samples).convert("RGBA")
        pr=page.rect; sx=pix.width/pr.width; sy=pix.height/pr.height
        self.verify_sx=sx; self.verify_sy=sy
        draw=ImageDraw.Draw(base,"RGBA")
        badge_font=self._load_font(9)
        STATUS_COL   = {"correct":(39,174,96,220),"incorrect":(231,76,60,220),"unknown":(52,152,219,180)}
        STATUS_BADGE = {"correct":(39,174,96,255),"incorrect":(231,76,60,255),"unknown":(52,152,219,255)}
        STATUS_ICON  = {"correct":"✓","incorrect":"✗","unknown":"?"}
        # ref triangle for this page
        refs=self.verify_orig_refs.get(pidx)
        if refs:
            sp=[(p[0]*sx,p[1]*sy) for p in refs]
            REF_COLS_PIL=[(240,180,41,255),(39,174,96,255),(231,76,60,255)]
            if len(sp)==3:
                for i in range(3): draw.line([sp[i],sp[(i+1)%3]],fill=(149,165,166,120),width=1)
            elif len(sp)==2: draw.line([sp[0],sp[1]],fill=(149,165,166,120),width=1)
            for i,(bx,by) in enumerate(sp):
                col=REF_COLS_PIL[i]; r=7
                draw.ellipse((bx-r,by-r,bx+r,by+r),outline=col,width=2)
                draw.line((bx-r-4,by,bx+r+4,by),fill=col,width=2)
                draw.line((bx,by-r-4,bx,by+r+4),fill=col,width=2)
        # markers
        for m in self.verify_markers:
            if m.get("page",0)!=pidx: continue
            mid=m.get("id",0)
            nx,ny,nw,nh=self._transform_rect(m["x"],m["y"],m["width"],m["height"],refs,None)
            x=int(round(nx*sx)); y=int(round(ny*sy))
            w=max(1,int(round(nw*sx))); h=max(1,int(round(nh*sy)))
            status=self.verify_statuses.get(mid,"unknown")
            col=STATUS_COL[status]; bcol=STATUS_BADGE[status]; icon=STATUS_ICON[status]
            is_sel=(mid==self.verify_selected_id)
            overlay=Image.new("RGBA",(w,h),(col[0],col[1],col[2],60))
            base.alpha_composite(overlay,(x,y))
            draw.rectangle((x,y,x+w,y+h),outline=bcol,width=4 if is_sel else 2)
            badge_txt=f"{icon}{mid}"; badge_w=max(28,len(badge_txt)*7+6)
            badge_y0=max(0,y-20)
            draw.rectangle((x,badge_y0,x+badge_w,badge_y0+20),fill=bcol,outline=bcol)
            if badge_font:
                tb=draw.textbbox((0,0),badge_txt,font=badge_font)
                tw2,th2=tb[2]-tb[0],tb[3]-tb[1]
                draw.text((x+(badge_w-tw2)/2,badge_y0+(20-th2)/2-tb[1]),badge_txt,font=badge_font,fill=(255,255,255,255))
            else:
                draw.text((x+2,badge_y0+2),badge_txt,fill=(255,255,255,255))
            if is_sel: draw.rectangle((x-3,y-3,x+w+3,y+h+3),outline=(255,200,0,255),width=2)
        self.verify_photo=ImageTk.PhotoImage(base.convert("RGB"))
        ww,hh=base.size
        self.verify_canvas.config(scrollregion=(0,0,ww,hh))
        self.verify_canvas.delete("all")
        self.verify_canvas.create_image(0,0,anchor=tk.NW,image=self.verify_photo)
        self.lbl_verify_page.config(text=f" {pidx+1} / {len(self.verify_doc)} ")

    def _update_verify_summary(self):
        total=len(self.verify_markers)
        c=sum(1 for s in self.verify_statuses.values() if s=="correct")
        w=sum(1 for s in self.verify_statuses.values() if s=="incorrect")
        u=sum(1 for s in self.verify_statuses.values() if s=="unknown")
        self.lbl_verify_summary.config(text=f"总计: {total} | ✓ 正确: {c} | ✗ 错误: {w} | ? 未核对: {u}")

    def _verify_export_results(self):
        if not self.verify_markers: messagebox.showwarning("提示","请先加载标记 JSON"); return
        save_path=filedialog.asksaveasfilename(title="保存核对结果",defaultextension=".xlsx",filetypes=[("Excel 文件","*.xlsx")])
        if not save_path: return
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="核对结果"
        ws.append(["序号","Excel内容","核对状态","备注"])
        for cell in ws[1]:
            cell.font=Font(bold=True,name="Arial"); cell.fill=PatternFill("solid",fgColor="D6EAF8")
            cell.alignment=Alignment(horizontal="center")
        ST={"correct":"✓ 正确","incorrect":"✗ 错误","unknown":"? 未核对"}
        SC={"correct":"C6EFCE","incorrect":"FFC7CE","unknown":"FFEB9C"}
        for m in sorted(self.verify_markers,key=lambda x:x.get("id",0)):
            mid=m.get("id",0); idx=mid-1
            content=self.verify_excel_data[idx] if 0<=idx<len(self.verify_excel_data) else ""
            status=self.verify_statuses.get(mid,"unknown")
            ws.append([mid,content,ST[status],self.verify_notes.get(mid,"")])
            row=ws.max_row
            ws.cell(row,3).fill=PatternFill("solid",fgColor=SC[status])
            ws.cell(row,3).alignment=Alignment(horizontal="center")
        ws.column_dimensions["A"].width=8; ws.column_dimensions["B"].width=50
        ws.column_dimensions["C"].width=14; ws.column_dimensions["D"].width=30
        wb.save(save_path); messagebox.showinfo("导出成功",f"核对结果已保存：\n{save_path}")


    # ═══════════════════════════════════════════════════════════════════════
    # SHARED FONT / TEXT UTILITIES
    # ═══════════════════════════════════════════════════════════════════════
    @classmethod
    def _font_candidates(cls):
        candidates=[]
        if UNICODE_FONT_PATH: candidates.append(UNICODE_FONT_PATH)
        system=platform.system()
        if system=="Windows":
            candidates.extend(["C:/Windows/Fonts/msyh.ttc","C:/Windows/Fonts/simhei.ttf",
                "C:/Windows/Fonts/seguisym.ttf","C:/Windows/Fonts/segoeuisym.ttf","C:/Windows/Fonts/seguiemj.ttf",
                "C:/Windows/Fonts/cambria.ttc","C:/Windows/Fonts/arialuni.ttf","C:/Windows/Fonts/segoeui.ttf",
                "C:/Windows/Fonts/arial.ttf","C:/Windows/Fonts/calibri.ttf"])
        elif system=="Darwin":
            candidates.extend(["/System/Library/Fonts/Apple Color Emoji.ttc","/System/Library/Fonts/Apple Symbols.ttf",
                "/Library/Fonts/Arial Unicode MS.ttf","/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
                "/System/Library/Fonts/Helvetica.ttc"])
        else:
            candidates.extend(["/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
                "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
                "/usr/share/fonts/truetype/noto/NotoSansSymbols2-Regular.ttf",
                "/usr/share/fonts/truetype/noto/NotoSansMath-Regular.ttf",
                "/usr/share/fonts/truetype/noto/NotoColorEmoji.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
                "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
                "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
                "/usr/share/fonts/truetype/stix/STIXGeneral.ttf"])
        seen,out=[],[]
        for p in candidates:
            if p and p not in seen and os.path.exists(p): seen.append(p); out.append(p)
        return out

    @classmethod
    def _font_pool(cls,size):
        pool={"regular":[],"cjk":[],"symbol":[],"emoji":[]}
        for p in cls._font_candidates():
            try: f=ImageFont.truetype(p,size=size)
            except: continue
            name=os.path.basename(p).lower()
            if "emoji" in name: pool["emoji"].append(f)
            elif any(x in name for x in ("sym","symbol","math","stix","cambria")): pool["symbol"].append(f)
            elif any(x in name for x in ("cjk","msyh","simhei","wqy","noto")): pool["cjk"].append(f)
            else: pool["regular"].append(f)
        for key in pool:
            if not pool[key]: pool[key]=[ImageFont.load_default()]
        return pool

    @staticmethod
    def _font_support_category(ch):
        cp=ord(ch); cat=unicodedata.category(ch)
        if ch in "\n\r\t": return "regular"
        if 0x1F000<=cp<=0x1FAFF or 0x2600<=cp<=0x27BF: return "emoji"
        if 0x2500<=cp<=0x259F or 0x2190<=cp<=0x22FF or 0x2300<=cp<=0x23FF: return "symbol"
        if 0x2000<=cp<=0x206F or 0x2070<=cp<=0x209F or 0x2100<=cp<=0x214F: return "symbol"
        if 0x3000<=cp<=0x303F or 0x3040<=cp<=0x30FF or 0x3400<=cp<=0x9FFF or 0xF900<=cp<=0xFAFF: return "cjk"
        if cat in {"So","Sm","Sk"}: return "symbol"
        return "regular"

    @classmethod
    def _pick_font_for_char(cls,pool,ch):
        key=cls._font_support_category(ch)
        for f in pool.get(key,[]): return f
        for k2 in ("regular","cjk","symbol","emoji"):
            if pool.get(k2): return pool[k2][0]
        return ImageFont.load_default()

    @staticmethod
    def _cell_display_text(cell):
        if cell is None or cell.value is None: return ""
        v=cell.value
        if isinstance(v,(datetime,date)):
            try:
                if isinstance(v,datetime):
                    if v.hour==0 and v.minute==0 and v.second==0 and v.microsecond==0: return v.strftime("%Y-%m-%d")
                    return v.strftime("%Y-%m-%d %H:%M:%S")
                return v.strftime("%Y-%m-%d")
            except: return str(v)
        try:
            if isinstance(v,float) and v.is_integer(): return str(int(v))
        except: pass
        return str(v)

    @classmethod
    def _wrap_text(cls,draw,text,pool,max_width):
        text="" if text is None else str(text)
        text=text.replace("\r\n","\n").replace("\r","\n"); lines=[]
        for para in text.split("\n"):
            if para=="": lines.append(""); continue
            cur=""; cur_w=0
            for ch in para:
                font=cls._pick_font_for_char(pool,ch); ct=" " if ch=="\t" else ch
                bbox=draw.textbbox((0,0),ct,font=font); cw=bbox[2]-bbox[0]
                if cur and cur_w+cw>max_width: lines.append(cur); cur=ch; cur_w=cw
                else: cur+=ch; cur_w+=cw
            if cur: lines.append(cur)
        return lines

    @classmethod
    def _draw_text_line(cls,draw,x,y,line,pool,max_width,fg):
        cx=x
        for ch in line:
            font=cls._pick_font_for_char(pool,ch); ct=" " if ch=="\t" else ch
            bbox=draw.textbbox((0,0),ct,font=font); cw=bbox[2]-bbox[0]
            if ch!="\r": draw.text((cx,y),ct,font=font,fill=fg)
            cx+=cw
            if cx-x>max_width+50: break

    @classmethod
    def _make_text_image(cls,text,box_w,box_h,bg=(214,234,248,255),fg=(26,37,47,255),fixed_font_size=None):
        box_w=max(12,int(box_w)); box_h=max(12,int(box_h))
        padding=4; max_text_w=max(4,box_w-padding*2); max_text_h=max(4,box_h-padding*2)
        probe=Image.new("RGBA",(max(20,box_w),max(20,box_h)),(0,0,0,0)); draw=ImageDraw.Draw(probe)
        def _measure(size):
            pool=cls._font_pool(size); lines=cls._wrap_text(draw,text,pool,max_text_w)
            gap=max(1,size//5); total_h=0; max_lw=0; metrics=[]
            for ln in lines:
                if ln=="":
                    f=cls._pick_font_for_char(pool,"A"); b=draw.textbbox((0,0),"A",font=f); lh=b[3]-b[1]; lw=0
                else:
                    tw2=0; lh=0
                    for ch in ln:
                        f=cls._pick_font_for_char(pool,ch); ct=" " if ch=="\t" else ch
                        b=draw.textbbox((0,0),ct,font=f); tw2+=b[2]-b[0]; lh=max(lh,b[3]-b[1])
                    lw=tw2; lh=max(lh,size)
                metrics.append((lw,lh)); max_lw=max(max_lw,lw); total_h+=lh
            total_h+=gap*max(0,len(lines)-1)
            return pool,lines,metrics,max_lw,total_h,gap
        def _render(pool,lines,metrics,box_w,box_h,gap):
            img=Image.new("RGBA",(box_w,box_h),bg); d=ImageDraw.Draw(img)
            y=max(0,(box_h-sum(m[1] for m in metrics)-gap*max(0,len(lines)-1))//2)
            for ln,(lw,lh) in zip(lines,metrics):
                x=padding if lw>=max_text_w else max(0,(box_w-lw)//2)
                cls._draw_text_line(d,x,y,ln,pool,max_text_w,fg); y+=lh+gap
            return img
        if fixed_font_size is not None:
            pool,lines,metrics,_,_,gap=_measure(max(4,int(fixed_font_size)))
            return _render(pool,lines,metrics,box_w,box_h,gap)
        for size in range(min(24,max(8,box_h//2)),5,-1):
            pool,lines,metrics,max_lw,total_h,gap=_measure(size)
            if max_lw<=max_text_w and total_h<=max_text_h:
                return _render(pool,lines,metrics,box_w,box_h,gap)
        pool,lines,metrics,_,_,gap=_measure(6)
        return _render(pool,lines,metrics,box_w,box_h,gap)

    @classmethod
    def _load_font(cls,size):
        for p in cls._font_candidates():
            try: return ImageFont.truetype(p,size=size)
            except: continue
        try: return ImageFont.load_default()
        except: return None

    @staticmethod
    def _draw_marker_on_canvas(canvas,sx,sy,sw,sh,mid,color):
        canvas.create_rectangle(sx,sy,sx+sw,sy+sh,outline=color,width=2,tags="marker")
        badge_w=max(22,len(str(mid))*7+6)
        canvas.create_rectangle(sx,sy-18,sx+badge_w,sy,fill=color,outline=color,tags="marker")
        canvas.create_text(sx+badge_w/2,sy-9,text=str(mid),fill="white",font=("Arial",9,"bold"),tags="marker")

    def __del__(self):
        for p in getattr(self,"_tmp_files",[]):
            try: os.unlink(p)
            except: pass


def main():
    root=tk.Tk()
    PDFMarkerApp(root)
    root.mainloop()

if __name__=="__main__":
    main()
